from flask import Blueprint, request, send_file, jsonify
from flask_login import login_required
import io
import pandas as pd

from database import (
    get_ledger_details,
    get_inventory_details,
    get_ledger_transactions,
    get_trial_balance_data,
    get_stock_movement_data,
    get_balance_sheet_data,
    get_profit_and_loss_data,
    get_closing_inventory_data,
    get_voucher_register_data,
)
from .models import parse_date, format_date
from . import get_db_connection

export_bp = Blueprint("export_bp", __name__)


@export_bp.route("/export_ledger")
@login_required
def export_ledger():
    try:
        ledgers = get_ledger_details()
        df = pd.DataFrame(
            ledgers,
            columns=[
                "Ledger Code",
                "Ledger Name",
                "Group Code",
                "Group Name",
                "Nature",
                "Opening Balance",
                "Opening Balance Type",
                "Opening Balance Date",
                "Closing Balance",
            ],
        )
        output = io.BytesIO()
        df.to_excel(output, index=False)
        output.seek(0)
        return send_file(
            output,
            download_name="ledger_details.xlsx",
            as_attachment=True,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        print(f"Error in export_ledger: {str(e)}")
        return jsonify({"success": False, "message": str(e)}), 500


@export_bp.route("/export_inventory")
@login_required
def export_inventory():
    try:
        inventory = get_inventory_details()
        from database import get_selling_price_map
        selling_price_map = get_selling_price_map()
        rows = [
            {
                "Item Code": i.get("item_code"),
                "Item Name": i.get("name"),
                "Group Code": i.get("stock_group_code"),
                "Group Name": i.get("group_name"),
                "Unit": i.get("unit_code"),
                "Selling Price": selling_price_map.get(i.get("item_code"), ""),
                "Opening Quantity": i.get("stock_quantity"),
                "VAT %": i.get("vat_rate"),
                "Opening Price (Cost)": i.get("opening_price"),
                "Location": i.get("opening_location_name"),
            }
            for i in inventory
        ]
        df = pd.DataFrame(
            rows,
            columns=[
                "Item Code",
                "Item Name",
                "Group Code",
                "Group Name",
                "Unit",
                "Selling Price",
                "Opening Quantity",
                "VAT %",
                "Opening Price (Cost)",
                "Location",
            ],
        )
        df["VAT %"] = df["VAT %"].apply(lambda x: float(x or 0))
        output = io.BytesIO()
        df.to_excel(output, index=False)
        output.seek(0)
        print(f"Exported {len(inventory)} inventory items with full columns")
        return send_file(
            output,
            download_name="inventory_details.xlsx",
            as_attachment=True,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        print(f"Error in export_inventory: {str(e)}")
        return jsonify({"success": False, "message": str(e)}), 500


@export_bp.route("/export_report/ledger_transactions")
@login_required
def export_ledger_transactions():
    ledger_name = request.args.get("ledger_name")
    from_date = parse_date(request.args.get("from_date"))
    to_date = parse_date(request.args.get("to_date"))
    try:
        transactions, _closing_balance = get_ledger_transactions(ledger_name, from_date, to_date)
        # Format dates
        formatted_transactions = []
        for t in transactions:
            # t is a dict: {'date': ..., 'voucher_number': ..., 'voucher_type': ..., 'narration': ..., 'debit': ..., 'credit': ..., 'balance': ...}
            formatted_transactions.append([
                t['voucher_number'],
                t['voucher_type'],
                format_date(t['date']),
                t['narration'],
                t['debit'],
                t['credit'],
                t['balance']
            ])
            
        df = pd.DataFrame(formatted_transactions, columns=["Voucher Number", "Voucher Type", "Date", "Narration", "Debit", "Credit", "Balance"])
        output = io.BytesIO()
        df.to_excel(output, index=False)
        output.seek(0)
        return send_file(output, download_name=f"{ledger_name}_transactions.xlsx", as_attachment=True)
    except Exception as e:
        print(f"Error in export_ledger_transactions: {str(e)}")
        return jsonify({"success": False, "message": str(e)}), 500


@export_bp.route("/export_report/trial_balance")
@login_required
def export_trial_balance():
    as_of_date = parse_date(request.args.get("as_of_date"))
    try:
        trial_balance, total_debit, total_credit = get_trial_balance_data(as_of_date)
        data = []
        for row in trial_balance:
            data.append([row["group_name"], row["ledger_name"], round(row["debit"], 2), round(row["credit"], 2)])
        data.append(["", "Total", round(total_debit, 2), round(total_credit, 2)])
        df = pd.DataFrame(data, columns=["Group", "Ledger Name", "Debit", "Credit"])
        output = io.BytesIO()
        df.to_excel(output, index=False)
        output.seek(0)
        return send_file(output, download_name="trial_balance.xlsx", as_attachment=True)
    except Exception as e:
        print(f"Error in export_trial_balance: {str(e)}")
        return jsonify({"success": False, "message": str(e)}), 500


@export_bp.route("/export_report/stock_movement")
@login_required
def export_stock_movement():
    item_name = request.args.get("item_name")
    from_date = parse_date(request.args.get("from_date"))
    to_date = parse_date(request.args.get("to_date"))
    try:
        movement = get_stock_movement_data(item_name, from_date, to_date)
        # movement tuples: (vn, dt, vt, qty_in, qty_out, running_qty, wap, running_val)
        
        formatted_movement = []
        for m in movement:
            fm = list(m)
            fm[1] = format_date(fm[1]) # Format date at index 1
            formatted_movement.append(fm)
            
        df = pd.DataFrame(formatted_movement, columns=[
            "Voucher Number", "Date", "Voucher Type",
            "Inward Qty", "Outward Qty",
            "Closing Qty", "WAP", "Closing Value", "Location"
        ])

        df = df[[
            "Date", "Voucher Type", "Voucher Number", "Location",
            "Inward Qty", "Outward Qty",
            "Closing Qty", "WAP", "Closing Value"
        ]]
        
        output = io.BytesIO()
        df.to_excel(output, index=False)
        output.seek(0)
        return send_file(output, download_name=f"{item_name}_stock_movement.xlsx", as_attachment=True)
    except Exception as e:
        print(f"Error in export_stock_movement: {str(e)}")
        return jsonify({"success": False, "message": str(e)}), 500


@export_bp.route("/export_report/balance_sheet")
@login_required
def export_balance_sheet():
    as_of_date = parse_date(request.args.get("as_of_date"))
    try:
        balance_sheet, total_assets, total_liabilities = get_balance_sheet_data(as_of_date)
        data = []
        
        # Assets
        data.append(["Assets", "", ""])
        for group_name, ledgers in balance_sheet["assets"].items():
            data.append([group_name, "", ""])
            for ledger in ledgers:
                data.append(["", ledger["ledger_name"], ledger["amount"]])
            group_total = round(sum(l['amount'] for l in ledgers), 2)
            data.append(["Total " + group_name, "", group_total])
            data.append([])
        data.append(["Total Assets", "", total_assets])
        
        data.append([])
        
        # Liabilities
        data.append(["Liabilities & Capital", "", ""])
        for group_name, ledgers in balance_sheet["liabilities"].items():
            data.append([group_name, "", ""])
            for ledger in ledgers:
                data.append(["", ledger["ledger_name"], ledger["amount"]])
            group_total = round(sum(l['amount'] for l in ledgers), 2)
            data.append(["Total " + group_name, "", group_total])
            data.append([])
        data.append(["Total Liabilities & Capital", "", total_liabilities])

        df = pd.DataFrame(data, columns=["Group", "Ledger Name", "Amount"])
        output = io.BytesIO()
        df.to_excel(output, index=False)
        output.seek(0)
        return send_file(output, download_name="balance_sheet.xlsx", as_attachment=True)
    except Exception as e:
        print(f"Error in export_balance_sheet: {str(e)}")
        return jsonify({"success": False, "message": str(e)}), 500


@export_bp.route("/export_report/profit_and_loss")
@login_required
def export_profit_and_loss():
    from_date = parse_date(request.args.get("from_date"))
    to_date = parse_date(request.args.get("to_date"))
    try:
        profit_and_loss, total_income, total_expenses, net_profit = get_profit_and_loss_data(from_date, to_date)
        data = []
        data.append(["Income", "", ""])
        for group_name, ledgers in profit_and_loss["income"].items():
            data.append([group_name, "", ""])
            for ledger in ledgers:
                data.append(["", ledger["ledger_name"], ledger["amount"]])
        data.append(["Total Income", "", round(total_income, 2)])
        data.append(["", "", ""])
        data.append(["Expenses", "", ""])
        for group_name, ledgers in profit_and_loss["expenses"].items():
            data.append([group_name, "", ""])
            for ledger in ledgers:
                data.append(["", ledger["ledger_name"], ledger["amount"]])
        data.append(["Total Expenses", "", round(total_expenses, 2)])
        data.append(["Net Profit/Loss", "", round(net_profit, 2)])

        df = pd.DataFrame(data, columns=["Category", "Ledger Name", "Amount"])
        output = io.BytesIO()
        df.to_excel(output, index=False)
        output.seek(0)
        return send_file(output, download_name="profit_and_loss.xlsx", as_attachment=True)
    except Exception as e:
        print(f"Error in export_profit_and_loss: {str(e)}")
        return jsonify({"success": False, "message": str(e)}), 500


@export_bp.route("/export_report/closing_inventory")
@login_required
def export_closing_inventory():
    as_of_date = parse_date(request.args.get("as_of_date"))
    try:
        closing_inventory, total_cost_amount = get_closing_inventory_data(as_of_date)
        df = pd.DataFrame(
            closing_inventory,
            columns=["item_code", "item_name", "group_name", "location_name", "quantity", "wap", "cost_amount"],
        )
        df.columns = ["Item Code", "Item Name", "Item Group", "Location", "Quantity", "WAP", "Cost Amount"]
        df.loc[len(df)] = ["Total", "", "", "", "", "", total_cost_amount]
        output = io.BytesIO()
        df.to_excel(output, index=False)
        output.seek(0)
        return send_file(output, download_name="closing_inventory.xlsx", as_attachment=True)
    except Exception as e:
        print(f"Error in export_closing_inventory: {str(e)}")
        return jsonify({"success": False, "message": str(e)}), 500


@export_bp.route("/export_report/voucher_register")
@login_required
def export_voucher_register():
    voucher_type = request.args.get("voucher_type")
    from_date = parse_date(request.args.get("from_date"))
    to_date = parse_date(request.args.get("to_date"))
    
    try:
        # get_voucher_register_data returns list of vouchers with items
        vouchers = get_voucher_register_data(voucher_type, from_date, to_date)
        
        # Flatten for Excel
        flattened_data = []
        
        for v in vouchers:
            # v keys: voucher_number, date, amount, narration, party_name, items, voucher_type
            v_date = format_date(v['date'])
            v_no = v['voucher_number']
            v_type = v['voucher_type']
            party = v['party_name']
            narration = v['narration']
            
            # Check if items exist and is a list
            if v.get('items') and isinstance(v['items'], list) and len(v['items']) > 0:
                for item in v['items']:
                    flattened_data.append([
                        v_date,
                        v_no,
                        v_type,
                        party,
                        item.get('name', ''),
                        item.get('qty', 0),
                        item.get('rate', 0),
                        item.get('amount', 0),
                        narration
                    ])
            else:
                # If no items (e.g. accounting voucher only?), still show row
                flattened_data.append([
                    v_date,
                    v_no,
                    v_type,
                    party,
                    "", # Item Name
                    0,  # Qty
                    0,  # Rate
                    v['amount'], # Amount
                    narration
                ])
                
        df = pd.DataFrame(flattened_data, columns=[
            "Date", "Voucher No", "Type", "Party", 
            "Item Name", "Quantity", "Rate", "Amount", "Narration"
        ])
        
        output = io.BytesIO()
        df.to_excel(output, index=False)
        output.seek(0)
        
        filename = f"{voucher_type}_Register.xlsx"
        return send_file(
            output, 
            download_name=filename, 
            as_attachment=True,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
    except Exception as e:
        print(f"Error in export_voucher_register: {str(e)}")
        return jsonify({"success": False, "message": str(e)}), 500


# Every chart openpyxl can build, keyed by the name the chat layer passes in.
# Value is (factory, kind) - "kind" tells _add_chart how to feed it data:
#   "category" - categories from a text column, one series per numeric column
#   "xy"       - first numeric column is X, the rest are Y series
#   "bubble"   - X, Y and bubble size from three numeric columns
CHART_SPECS = {
    # --- Pie family: one series only ---
    "pie": ("PieChart", "category"),
    "pie3d": ("PieChart3D", "category"),
    "doughnut": ("DoughnutChart", "category"),
    "pie_of_pie": ("ProjectedPieChart", "category"),
    "bar_of_pie": ("ProjectedPieChart", "category"),
    # --- Bar / column ---
    "bar": ("BarChart", "category"),
    "hbar": ("BarChart", "category"),
    "bar3d": ("BarChart3D", "category"),
    "stacked_bar": ("BarChart", "category"),
    "percent_stacked_bar": ("BarChart", "category"),
    # --- Line ---
    "line": ("LineChart", "category"),
    "line3d": ("LineChart3D", "category"),
    "stacked_line": ("LineChart", "category"),
    # --- Area ---
    "area": ("AreaChart", "category"),
    "area3d": ("AreaChart3D", "category"),
    "stacked_area": ("AreaChart", "category"),
    # --- Radar ---
    "radar": ("RadarChart", "category"),
    "filled_radar": ("RadarChart", "category"),
    # --- Surface ---
    "surface": ("SurfaceChart", "category"),
    "surface3d": ("SurfaceChart3D", "category"),
    # --- Stock: needs high/low/close columns ---
    "stock": ("StockChart", "category"),
    # --- XY ---
    "scatter": ("ScatterChart", "xy"),
    "bubble": ("BubbleChart", "bubble"),
}

# Charts that can only ever show a single series.
_SINGLE_SERIES = {"pie", "pie3d", "doughnut", "pie_of_pie", "bar_of_pie"}

# Minimum numeric columns each shape needs before it can be drawn at all.
_MIN_NUMERIC = {"xy": 2, "bubble": 3}


def _configure(chart, chart_type):
    """Apply the options that distinguish variants sharing one openpyxl class."""
    from openpyxl.chart.label import DataLabelList

    if chart_type in ("hbar", "stacked_bar", "percent_stacked_bar", "bar"):
        chart.type = "bar" if chart_type == "hbar" else "col"
    if chart_type == "stacked_bar":
        chart.grouping = "stacked"
        chart.overlap = 100
    elif chart_type == "percent_stacked_bar":
        chart.grouping = "percentStacked"
        chart.overlap = 100
    elif chart_type == "stacked_line":
        chart.grouping = "stacked"
    elif chart_type == "stacked_area":
        chart.grouping = "stacked"
    elif chart_type == "filled_radar":
        chart.type = "filled"
    elif chart_type == "radar":
        chart.type = "marker"
    elif chart_type == "bar_of_pie":
        chart.type = "bar"
    elif chart_type == "pie_of_pie":
        chart.type = "pie"

    if chart_type in _SINGLE_SERIES:
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showPercent = True


def _add_chart(worksheet, df, chart_type):
    """Draw the requested chart next to the data on `worksheet`.

    Silently does nothing when the shape of the data cannot carry the chart -
    the spreadsheet itself is the answer, the chart is a bonus.
    """
    from openpyxl import chart as _chart
    from openpyxl.chart import Reference, Series
    from openpyxl.utils import get_column_letter

    if df.empty:
        return

    spec = CHART_SPECS.get(chart_type)
    if spec is None:
        chart_type, spec = "bar", CHART_SPECS["bar"]
    factory_name, kind = spec

    numeric = [i for i, c in enumerate(df.columns) if pd.api.types.is_numeric_dtype(df[c])]
    labels = [i for i, c in enumerate(df.columns) if i not in numeric]
    if not numeric:
        return

    # Not enough numeric columns for an XY-style chart - a category chart of
    # the same numbers is still useful, so fall back rather than skip.
    if len(numeric) < _MIN_NUMERIC.get(kind, 0):
        chart_type, factory_name, kind = "bar", "BarChart", "category"
    # A stock chart without high/low/close would render as an empty frame.
    if chart_type == "stock" and len(numeric) < 3:
        chart_type, factory_name = "line", "LineChart"
    # A surface chart of a single series draws a meaningless ribbon.
    if chart_type in ("surface", "surface3d") and len(numeric) < 2:
        chart_type, factory_name = "bar", "BarChart"

    chart = getattr(_chart, factory_name)()
    _configure(chart, chart_type)
    last_row = len(df) + 1  # +1 for the header row

    if kind == "category":
        # An all-numeric result (e.g. year | total) still charts fine - the
        # leading number reads as the axis label, the rest as the values.
        if labels:
            label_col, value_pool = labels[0], numeric
        elif len(numeric) >= 2:
            label_col, value_pool = numeric[0], numeric[1:]
        else:
            return
        value_cols = value_pool[:1] if chart_type in _SINGLE_SERIES else value_pool
        for col in value_cols:
            chart.add_data(
                Reference(worksheet, min_col=col + 1, min_row=1, max_row=last_row),
                titles_from_data=True,
            )
        chart.set_categories(
            Reference(worksheet, min_col=label_col + 1, min_row=2, max_row=last_row)
        )
        head = df.columns[value_cols[0]] if len(value_cols) == 1 else "Values"
        chart.title = str(head) + " by " + str(df.columns[label_col])
    elif kind == "xy":
        xvalues = Reference(worksheet, min_col=numeric[0] + 1, min_row=2, max_row=last_row)
        for col in numeric[1:]:
            values = Reference(worksheet, min_col=col + 1, min_row=1, max_row=last_row)
            chart.series.append(Series(values, xvalues, title_from_data=True))
        chart.x_axis.title = str(df.columns[numeric[0]])
        chart.title = str(df.columns[numeric[1]]) + " vs " + str(df.columns[numeric[0]])
    else:  # bubble
        xvalues = Reference(worksheet, min_col=numeric[0] + 1, min_row=2, max_row=last_row)
        yvalues = Reference(worksheet, min_col=numeric[1] + 1, min_row=2, max_row=last_row)
        size = Reference(worksheet, min_col=numeric[2] + 1, min_row=2, max_row=last_row)
        series = Series(yvalues, xvalues, zvalues=size, title=str(df.columns[numeric[1]]))
        chart.series.append(series)
        chart.title = str(df.columns[numeric[1]]) + " vs " + str(df.columns[numeric[0]])

    chart.height, chart.width = 9, 15
    worksheet.add_chart(chart, get_column_letter(len(df.columns) + 2) + "2")


def _chat_result_pdf(result):
    """A landscape PDF of the result table, or None if reportlab is missing."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                        Paragraph, Spacer)
    except ImportError:
        return None

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            leftMargin=12 * mm, rightMargin=12 * mm,
                            topMargin=12 * mm, bottomMargin=12 * mm)
    styles = getSampleStyleSheet()
    cell = styles["BodyText"].clone("cell")
    cell.fontSize = 7
    cell.leading = 9

    def text(value):
        if value is None:
            return ""
        if isinstance(value, float):
            return f"{value:,.2f}"
        return str(value)

    header = [Paragraph(f"<b>{c}</b>", cell) for c in result["columns"]]
    body = [[Paragraph(text(v), cell) for v in row] for row in result["rows"]]

    table = Table([header] + body, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f1f5f9")),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.white, colors.HexColor("#fafafa")]),
    ]))

    doc.build([Paragraph(f"<b>{result['title']}</b>", styles["Heading2"]),
               Spacer(1, 6), table])
    buffer.seek(0)
    return buffer


@export_bp.route("/export_chat_result")
@login_required
def export_chat_result():
    """Download a tabular answer the chat produced, as .xlsx, .csv or .pdf."""
    from .chat_export_store import load

    result = load(request.args.get("token"))
    if not result:
        return jsonify({
            "success": False,
            "message": "That result is no longer available. Ask the question again.",
        }), 404

    chart_type = (request.args.get("chart") or "").strip().lower()
    fmt = (request.args.get("format") or "xlsx").strip().lower()

    if fmt == "csv":
        df = pd.DataFrame(result["rows"], columns=result["columns"])
        buffer = io.BytesIO(df.to_csv(index=False).encode("utf-8-sig"))
        return send_file(buffer, download_name=result["title"] + ".csv",
                         as_attachment=True, mimetype="text/csv")

    if fmt == "pdf":
        pdf = _chat_result_pdf(result)
        if pdf is None:
            return jsonify({
                "success": False,
                "message": "PDF export needs the reportlab package. "
                           "Download the Excel or CSV file instead.",
            }), 501
        return send_file(pdf, download_name=result["title"] + ".pdf",
                         as_attachment=True, mimetype="application/pdf")

    try:
        df = pd.DataFrame(result["rows"], columns=result["columns"])
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Data")
            if chart_type:
                _add_chart(writer.sheets["Data"], df, chart_type)
        output.seek(0)
        return send_file(
            output,
            download_name=result["title"] + ".xlsx",
            as_attachment=True,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as e:
        print("Error in export_chat_result: " + str(e))
        return jsonify({"success": False, "message": str(e)}), 500


@export_bp.route("/export_report/statement_of_account")
@login_required
def export_statement_of_account():
    """The statement of account for one party, as .xlsx."""
    from database.soa_db import get_statement_of_account

    ledger_name = request.args.get("ledger") or ""
    as_of = request.args.get("as_of") or None
    include_settled = request.args.get("include_settled") == "1"

    statement = get_statement_of_account(ledger_name, as_of_date=as_of,
                                         include_settled=include_settled)
    if not statement:
        return jsonify({"success": False,
                        "message": "Select a customer or supplier first."}), 400

    columns = ["Date", "Invoice / Voucher", "Type", "Reference", "Due Date",
               "Days Overdue", "Original Amount", "Matched Amount",
               "Remaining Amount", "Status"]
    rows = [[r["date"], r["voucher_number"], r["voucher_type"], r["reference"],
             r["due_date"], r["days_overdue"], r["original_amount"],
             r["matched_amount"], r["remaining_amount"], r["status"]]
            for r in statement["rows"]]
    frame = pd.DataFrame(rows, columns=columns)

    totals = statement["totals"]
    summary = pd.DataFrame([
        ["Party", statement["ledger_name"]],
        ["Type", statement["party_kind"]],
        ["Group", statement["group_name"]],
        ["As of", statement["as_of_date"]],
        ["Outstanding invoices", len(statement["rows"])],
        ["Original invoiced", totals["original"]],
        ["Matched / settled", totals["matched"]],
        ["Remaining on invoices", totals["remaining"]],
        ["Payments not yet matched", totals["unallocated"]],
        ["Net outstanding", totals["net_outstanding"]],
        ["Ledger balance", statement["ledger_balance"]],
        ["Ties to the ledger", "Yes" if statement["reconciles"] else
         f"No - out by {statement['difference']:,.2f}"],
    ], columns=["Item", "Value"])

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary.to_excel(writer, index=False, sheet_name="Summary")
        frame.to_excel(writer, index=False, sheet_name="Outstanding Invoices")
        if statement["unallocated"]:
            pd.DataFrame(
                [[u["date"], u["voucher_number"], u["voucher_type"],
                  u["description"], u["amount"]] for u in statement["unallocated"]],
                columns=["Date", "Voucher", "Type", "Description",
                         "Unmatched Amount"]
            ).to_excel(writer, index=False, sheet_name="Unmatched Payments")

        for sheet_name, df in (("Summary", summary),
                               ("Outstanding Invoices", frame)):
            sheet = writer.sheets[sheet_name]
            for index, column in enumerate(df.columns, start=1):
                widest = max([len(str(column))] +
                             [len(str(v)) for v in df[column].head(200)] or [0])
                sheet.column_dimensions[
                    sheet.cell(row=1, column=index).column_letter
                ].width = min(max(widest + 2, 12), 46)
    output.seek(0)

    safe = "".join(c for c in statement["ledger_name"] if c not in '\/:*?"<>|')[:60]
    return send_file(
        output, download_name=f"SOA - {safe}.xlsx", as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@export_bp.route("/export_report/statement_of_account_pdf")
@login_required
def export_statement_of_account_pdf():
    """A printable statement to send to the customer or supplier."""
    from database.soa_db import get_statement_of_account
    from database.company_db import get_company_settings
    from .soa_pdf import build_statement_pdf

    ledger_name = request.args.get("ledger") or ""
    as_of = request.args.get("as_of") or None
    include_settled = request.args.get("include_settled") == "1"

    statement = get_statement_of_account(ledger_name, as_of_date=as_of,
                                         include_settled=include_settled)
    if not statement:
        return jsonify({"success": False,
                        "message": "Select a customer or supplier first."}), 400

    try:
        company = get_company_settings() or {}
    except Exception:
        company = {}

    try:
        pdf = build_statement_pdf(statement, company)
    except ImportError:
        return jsonify({"success": False,
                        "message": "PDF export needs the reportlab package."}), 501

    safe = "".join(c for c in statement["ledger_name"] if c not in '\/:*?"<>|')[:60]
    return send_file(pdf, download_name=f"Statement - {safe}.pdf",
                     as_attachment=True, mimetype="application/pdf")
